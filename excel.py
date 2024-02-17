from pydantic import BaseModel
from openpyxl import Workbook, load_workbook

# Pydantic model for the data to be received
class ExcelData(BaseModel):
    name: str
    email: str
    feedback: str

# Function to append data to an Excel sheet
def append_to_excel(data: ExcelData):
    # Load existing workbook or create a new one if not exists
    try:
        wb = load_workbook("feedback.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    # Select active worksheet (the first one by default)
    ws = wb.active

    # Append data to the next available row
    row = (data.name, data.email, data.feedback)
    ws.append(row)

    # Save the workbook
    wb.save("feedback.xlsx")


